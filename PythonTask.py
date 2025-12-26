import os
import pandas as pd
import openpyxl
import xlwings as xw
import json
import sys
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime
from concurrent_log_handler import ConcurrentRotatingFileHandler

# --- CUSTOM TIMESTAMPED ROTATING HANDLER ---
class TimestampedConcurrentRotatingFileHandler(ConcurrentRotatingFileHandler):
    def doRollover(self):
        if self.stream:
            self.stream.close()
            self.stream = None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dirname, basename = os.path.split(self.baseFilename)
        new_logname = os.path.join(dirname, f"master_{timestamp}.log")

        if os.path.exists(self.baseFilename):
            os.rename(self.baseFilename, new_logname)

        # All timestamped logs will be kept

        self.stream = self._open()

# --- CONFIG DECLARATION BLOCK ---
script_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(script_dir, 'config.json')
log_path = os.path.join(script_dir,'logs','Master.log')
with open(config_path, 'r') as f:
    config = json.load(f)
raw_file_source = config['raw_file_source']
bp_process_dest = config['bp_process_dest']
serum_process_dest = config['serum_process_dest']
bp_copy_source = config['bp_copy_source']
serum_copy_source = config['serum_copy_source']
bp_copy_dest = config['bp_copy_dest']
serum_copy_dest = config['serum_copy_dest']
interval_minutes = config['interval_minutes']
record_process = config['record_process']
log_folder = config['log_folder']
# --- END CONFIG DECLARATION BLOCK ---

# --- LOGGER SETUP BLOCK ---
formatter = logging.Formatter('[%(asctime)s] [%(name)s][%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
blood_logger = logging.getLogger("BP")
blood_handler = TimestampedConcurrentRotatingFileHandler(
    log_path, maxBytes=20*1024*1024, encoding='utf-8'
)
blood_stream = logging.StreamHandler(sys.stdout)
blood_handler.setFormatter(formatter)
blood_stream.setFormatter(formatter)
blood_logger.setLevel(logging.INFO)
blood_logger.addHandler(blood_handler)
blood_logger.addHandler(blood_stream)
blood_logger.propagate = False
serum_logger = logging.getLogger("SERUM")
serum_handler = TimestampedConcurrentRotatingFileHandler(
    log_path, maxBytes=20*1024*1024, encoding='utf-8'
)
serum_stream = logging.StreamHandler(sys.stdout)
serum_handler.setFormatter(formatter)
serum_stream.setFormatter(formatter)
serum_logger.setLevel(logging.INFO)
serum_logger.addHandler(serum_handler)
serum_logger.addHandler(serum_stream)
serum_logger.propagate = False
# --- END LOGGER SETUP BLOCK ---

# --- Open and read value in record_process xlsx file ---
def load_or_create_record(record_path):
    try:
        if os.path.exists(record_path):
            record_wb = load_workbook(record_path)
            record_ws = record_wb.active
        else:
            record_wb = Workbook()
            record_ws = record_wb.active
            record_ws.title = "Sheet1"
            record_ws.append(["Processed Files"])
            record_wb.save(record_path)
        processed_files = {cell.value for cell in record_ws['A'] if cell.value}
        return record_wb, record_ws, processed_files
    except Exception as e:
        blood_logger.error(f"Error loading/creating record workbook: {e}")
        sys.exit(1)

def update_record(file_name, record_ws, record_wb, record_path):
    record_ws.append([file_name])
    record_wb.save(record_path)

# --- Function to apply sensitivity label ---
def apply_sensitivity_label(output_path, logger):
    try:
        xl_app = xw.App(visible=False)
        xl_book = xl_app.books.open(output_path)
        try:
            label_info = xl_book.api.SensitivityLabel.CreateLabelInfo()
            label_info.AssignmentMethod = 2
            label_info.LabelId = "f48041ff-f5de-4583-8841-e2a1851ee5d2"
            label_info.LabelName = "Confidential"
            label_info.SiteId = "771c9c47-7f24-44dc-958e-34f8713a8394t"
            xl_book.api.SensitivityLabel.SetLabel(label_info, label_info)
        except Exception as e:
            logger.warning(f"Failed to apply sensitivity label to {output_path}: {e}")
        xl_book.save()
        xl_book.close()
        xl_app.quit()
    except Exception as e:
        logger.error(f"Error applying sensitivity label to {output_path}: {e}")

# --- first type of file format processing ---
def format_serum_excel(ws):
    ws["E1"] = "Weight (kg)"
    ws["G1"] = "Total Weight (kg)"
    ws["J1"] = "Total Litres Processed"
    for col in ws.columns:
        for cell in col:
            col_letter = cell.column_letter
            if col_letter == 'D':
                cell.number_format = "0"
            elif col_letter in ['E', 'G']:
                cell.number_format = "0.000"
            elif col_letter in ['F', 'H', 'I']:
                cell.number_format = "0.00"
            else:
                cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="left")
    for col in ['B', 'C']:
        ws.column_dimensions[col].auto_size = True
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 11
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border = Border(bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'), top=Side(style='thin'))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True)
    ws.protection.sheet = True
    ws.protection.set_password("password")
    ws.protection.enable_selection = 'UnlockedCells'

# --- Second type of file format processing ---
def format_bp_excel(ws):
    ws.insert_cols(7)
    ws['G1'] = "SFF Net Weight (kg)"
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.protection = Protection(locked=False)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    formula = 'AND(P2<>"", ABS(P2 - SUM(F2:F1087)) > 0.2)'
    rule = FormulaRule(formula=[formula], fill=red_fill)
    ws.conditional_formatting.add('P2', rule)
    highlight_fill = PatternFill(start_color="EC4F28", end_color="EC4F28", fill_type="solid")
    rule2 = FormulaRule(formula=['AND(G2<>"", M2="Accept")'], fill=highlight_fill)
    ws.conditional_formatting.add(f'G2:G{ws.max_row}', rule2)
    ws.insert_cols(9)
    ws['I1'] = "Volume (L)"
    for row in range(2, ws.max_row + 1):
        formula = f'=IF(G{row}="", H{row}, G{row}*0.959)'
        ws[f'I{row}'] = formula
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='left')
    for col_letter in ['D']:
        for cell in ws[col_letter]:
            cell.number_format = '0'
    for col_letter in ['E', 'F', 'G', 'J', 'K']:
        for cell in ws[col_letter]:
            cell.number_format = '0.000'
    for col_letter in ['H', 'I', 'L', 'N', 'O', 'Q']:
        for cell in ws[col_letter]:
            cell.number_format = '0.00'
    ws['E1'] = "Gross Weight (kg)"
    ws['F1'] = "Net Weight (kg)"
    ws['G1'] = "SFF Net Weight (kg)"
    ws['H1'] = "Volume Recorded (L)"
    ws['I1'] = "Volume (L)"
    ws['J1'] = "Total Gross Weight (kg)"
    ws['K1'] = "Total Net Weight (kg)"
    ws['L1'] = "Total Volume (L)"
    ws['M1'] = "Accept / Reject"
    ws['N1'] = "Total Litres Rejected"
    ws['O1'] = "Total Litres Accepted"
    ws['P1'] = "Total SFF Net Weight (kg)"
    ws['Q1'] = "SFF Litres"
    ws['N2'] = '=SUMIF(M:M, "Reject", I:I)'
    ws['O2'] = '=IF(ISBLANK(P2), SUMIF(M2:M1048576, "Accept", I2:I1048576), Q2-N2)'
    ws['Q2'] = "=P2*0.959"
    dv = DataValidation(type="list", formula1='"Accept,Reject"', showDropDown=False, allowBlank=False,
                        showErrorMessage=True, errorTitle="Invalid Entry",
                        error="Please select either Accept or Reject from the dropdown list")
    ws.add_data_validation(dv)
    dv.add("M2:M1048")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['O2'].fill = yellow_fill
    for cell in ws['K']:
        cell.fill = yellow_fill
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'), top=Side(style='thin'))
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 30
    column_widths = {
        'A': 29, 'B': 23, 'C': 10.5, 'D': 9, 'E': 12, 'F': 11, 'G': 12,
        'H': 13, 'I': 12, 'J': 12, 'K': 12, 'L': 11, 'M': 11, 'N': 11, 'O': 13, 'P': 13
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    for row in ws.iter_rows(min_row=2, min_col=13, max_col=13):
        for cell in row:
            cell.protection = Protection(locked=False)
    ws['P2'].protection = Protection(locked=False)
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.protection = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.set_password("password")
    ws.protection.enable_selection = 'UnlockedCells'

# --- first type of file apply password and sensitivity label, processing with the format_serum_excel ---
def process_serum_file(file_name, csv_path):
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        serum_logger.error(f"Error reading {file_name}: {e}")
        return False
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        output_path = os.path.join(serum_process_dest, file_name.replace(".csv", ".xlsx"))
        wb.save(output_path)
        apply_sensitivity_label(output_path, serum_logger)
        try:
            xl_app = xw.App(visible=False)
            xl_book = xl_app.books.open(output_path)
            xl_sheet = xl_book.sheets[0]
            xl_sheet.range("J2").formula = '=SUM(MAXIFS(F2:F1048576,D2:D1048576,UNIQUE(D2:D1048576)))'
            xl_book.save()
            xl_book.close()
            xl_app.quit()
        except Exception as e:
            serum_logger.warning(f"Failed to insert formula in {output_path}: {e}")
        wb = load_workbook(output_path)
        ws = wb.active
        format_serum_excel(ws)
        wb.save(output_path)
        serum_logger.info(f"First Type of file processed: {file_name} -> {output_path}")
        return True
    except Exception as e:
        serum_logger.error(f"Error processing serum file {file_name}: {e}")
        return False

# --- second type of file apply password and sensitivity label, processing with the format_bp_excel ---
def process_bp_file(file_name, csv_path):
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        blood_logger.error(f"Error reading {file_name}: {e}")
        return False
    try:
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        output_path = os.path.join(bp_process_dest, file_name.replace(".csv", ".xlsx"))
        wb.save(output_path)
        apply_sensitivity_label(output_path, blood_logger)
        wb = load_workbook(output_path)
        ws = wb.active
        format_bp_excel(ws)
        wb.save(output_path)
        blood_logger.info(f"Second Type of file processed: {file_name} -> {output_path}")
        return True
    except Exception as e:
        blood_logger.error(f"Error processing BP file {file_name}: {e}")
        return False

def main():
    record_wb, record_ws, processed_files = load_or_create_record(record_process)
    new_files_processed = 0
    for file_name in os.listdir(raw_file_source):
        csv_path = os.path.join(raw_file_source, file_name)
        if file_name.startswith("F") and file_name.endswith(".csv") and file_name not in processed_files:
            if process_serum_file(file_name, csv_path):
                update_record(file_name, record_ws, record_wb, record_process)
                new_files_processed += 1
        elif file_name.endswith(".csv") and "NZL" in file_name and file_name not in processed_files:
            if process_bp_file(file_name, csv_path):
                update_record(file_name, record_ws, record_wb, record_process)
                new_files_processed += 1
    if new_files_processed == 0:
        blood_logger.info("Nothing new to process.")

if __name__ == "__main__":
    main()